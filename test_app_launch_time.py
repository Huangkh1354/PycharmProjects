import os
import random
import subprocess
import time
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

#Some adb command
cmd_find_packages = ['adb', 'shell', 'cmd', 'package', 'query-activities',
             '--brief', '-a', 'android.intent.action.MAIN',
             '-c', 'android.intent.category.LAUNCHER']
cmd_return_launcher = ["adb", "shell", "am", "start",
        "-a", "android.intent.action.MAIN",
        "-c", "android.intent.category.HOME"]
cmd_get_device = ['adb', 'shell', 'getprop', '|', 'grep', 'ro.product.device']
cmd_kill_app = "adb shell am force-stop "
cmd_start_app = "adb shell am start -W "

global_packages = []
launch_time_test_result = []

def main():
    print(f'测试设备为: {get_device_name()}')
    get_launcher_apps()
    for app in global_packages:
        test_app_launch_time(app, True)

    print(launch_time_test_result)
    json_to_excel_pandas()
    print("All done!")
    input("按任意键退出...")

def get_launcher_apps():
    #获取已安装的app列表，筛选有MAIN Activity的app
    global global_packages
    res = run_adb_command(cmd_find_packages)
    if res['success']:
        global_packages = [
            line.strip()
            for line in res['output'].splitlines()
            if "/" in line and not line.strip().startswith((' ', '\t'))  # 排除缩进行
        ]
        print("\n待测试包名:")
        print('\n'.join(global_packages))
        print(f"\n总共获取到 {len(global_packages)} 个应用")
    else:
        global_packages = []
        print(f"获取应用列表错误：{res['output']}，code: {res['error']}")
        exit()

def test_app_launch_time(package, should_kill):
    #should_kill: true, launch state is COLD
    #             false, launch state is HOT or WARM
    package_name, activity_name = package.split('/')
    if should_kill:
        run_adb_command(cmd_kill_app + package_name)
        time.sleep(1)
    res = run_adb_command(cmd_start_app + package)
    if res['success']:
        print(res['output'])
        launch_time_test_result.append(format_json(res['output']))
        run_adb_command(cmd_return_launcher)
        time.sleep(1)
        if "COLD" in res['output']:
            test_app_launch_time(package, False)
            time.sleep(1)
        else:
            print("This app test done,kill!\n")
            run_adb_command(cmd_kill_app + package_name)
    else:
        error_json = {
            'package': package_name,
            'error': res['error'],
        }
        launch_time_test_result.append(error_json)

def run_adb_command(command, adb_path="adb"):
    """
    执行ADB命令并返回结果

    参数:
        command (str/list): 要执行的ADB命令（可以是字符串或列表）
        adb_path (str): adb可执行文件路径（默认使用系统PATH中的adb）

    返回:
        dict: {
            'success': bool,  # 是否成功执行
            'output': str,   # 命令输出(stdout)
            'error': str,     # 错误信息(stderr)
            'returncode': int # 返回码
        }
    """
    # 如果传入的是字符串，分割成列表
    if isinstance(command, str):
        command = command.split()

    # 确保命令以adb开头
    if command[0] != adb_path:
        command.insert(0, adb_path)

    try:
        result = subprocess.run(
            command,
            capture_output=True,
            text=True,
            check=False  # 不自动抛出异常，我们自己处理返回码
        )

        return {
            'success': result.returncode == 0,
            'output': result.stdout.strip(),
            'error': result.stderr.strip(),
            'returncode': result.returncode
        }

    except Exception as e:
        return {
            'success': False,
            'output': '',
            'error': str(e),
            'returncode': -1
        }

def format_json(raw_data):
    # 找到"Status:"开始的位置
    start_index = raw_data.find("Status:")
    if start_index == -1:
        print("未找到Status字段，打开应用失败")
        start_index = raw_data.find("Starting:")

    # 只处理从Status开始的部分
    relevant_data = raw_data[start_index:]

    # 转换为字典
    result = {}
    for line in relevant_data.splitlines():
        if ':' in line:
            key, value = line.split(':', 1)  # 只分割第一个冒号
            key = key.strip()
            value = value.strip()

            # 数值转换
            if key in ['TotalTime', 'WaitTime']:
                try:
                    result[key] = int(value)
                except ValueError:
                    result[key] = value
            else:
                result[key] = value
    if "Error" in result:
        result['Activity'] = result.pop('Starting', '').split("cmp=")[1].split("}")[0].strip()

    return result

def get_device_name():
    res = run_adb_command(cmd_get_device)
    if res['success']:
        return res['output'].split('[')[2].split(']')[0]
    else:
        return ''

def json_to_excel_pandas(
        output_file=f"launch_time_result_{get_device_name()}_{datetime.now().strftime("%Y_%m_%d_%H%M%S")}.xlsx"):
    #使用时间戳来命名文件，基本不会重名
    #以防万一加入随机码
    try:
        if os.path.exists(output_file):
            os.remove(output_file)
    except Exception as e:
        print(f"删除旧文件失败: {e}, 给文件名添加随机码")
        random_digits = ''.join(random.choices('0123456789', k=5))
        output_file=f"launch_time_result_{get_device_name()}_{datetime.now().strftime("%Y_%m_%d_%H%M%S")}_{random_digits}.xlsx"

    wb = Workbook()
    ws = wb.active

    # 获取所有列名
    all_keys = sorted(
        {key for item in launch_time_test_result for key in item.keys()},
        key=lambda x: (x == 'Error', x)
    )

    # 写入带样式的表头
    header_font = Font(bold=True, color="000000")
    ws.append(all_keys)
    for cell in ws[1]:
        cell.font = header_font

    # 写入数据
    for item in launch_time_test_result:
        ws.append([item.get(key, "") for key in all_keys])

    # 自动调整列宽
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

    wb.save(output_file)
    abs_path = os.path.abspath(output_file)
    print(f"测试结果文件已生成：{abs_path}")
    return True



# 按装订区域中的绿色按钮以运行脚本。
if __name__ == '__main__':
    main()
